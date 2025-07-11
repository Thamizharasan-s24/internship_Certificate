# app.py
from flask import Flask, render_template, request, send_file, redirect, url_for
from PIL import Image, ImageDraw, ImageFont
from datetime import datetime
import io
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
preview_cache = {}

# Font paths (must exist in static/fonts/)
FONT_DIR = os.path.join(app.root_path, 'static', 'fonts')
REGULAR_FONT_PATH = os.path.join(FONT_DIR, "times new roman.ttf")
BOLD_FONT_PATH = os.path.join(FONT_DIR, "times new roman bold.ttf")


@app.route('/')
def index():
    return render_template("index.html")

# ...existing code...
@app.route('/generate', methods=['POST'])
def generate():
    try:
        name = request.form['name']
        id_no = request.form['id_no']
        department = request.form['department']
        university = request.form['university']
        role = request.form['role']
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').strftime('%d.%m.%Y')
        end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').strftime('%d.%m.%Y')
        issue_date = datetime.strptime(request.form['issue_date'], '%Y-%m-%d').strftime('%d.%m.%Y')
        year = request.form['reg_year']
        number = request.form['reg_number'].zfill(3)
        reg_no = f"INT/{year}/{number}"
        file_format = request.form['file_format']

        # Load template image directly from static/templates/
        TEMPLATE_PATH = os.path.join(app.root_path, 'static', 'templates', 'DTS Internship Certificate.png')
        image = Image.open(TEMPLATE_PATH).convert("RGB")
        draw = ImageDraw.Draw(image)

        # ...existing code...

        # Load fonts
        try:
            regular_font = ImageFont.truetype(REGULAR_FONT_PATH, 30)
            bold_font = ImageFont.truetype(BOLD_FONT_PATH, 30)
        except OSError:
            return "Font files not found. Please check static/fonts/ path."

        # Header info
        draw.text((1050, 370), f"Reg   : {reg_no}", fill="black", font=regular_font)
        draw.text((1050, 400), f"Place : Bengaluru", fill="black", font=regular_font)
        draw.text((1050, 430), f"Date  : {issue_date}", fill="black", font=regular_font)

        # Paragraph parts
        part1 = [
            ("This is to certify that", regular_font),
            (name, bold_font),
            (", Reg No. ", regular_font),
            (id_no, bold_font),
            (", a student of ", regular_font),
            (department, bold_font),
            (", ", regular_font), 
            (university, bold_font),
            (", has successfully completed an internship at our organization in the role of ", regular_font),
            (role, bold_font),
            (".", regular_font)
        ]

        part2 = [
            ("The internship period was from ", regular_font),
            (start_date, bold_font),
            (" to ", regular_font),
            (end_date, bold_font),
            (", during which the student worked under my supervision and actively participated in various tasks and learning activities. ", regular_font),
            ("Throughout the internship, the student exhibited a strong commitment to learning, and demonstrated good understanding of Python programming fundamentals and their practical applications.", regular_font)
        ]

        part3 = [
            ("We appreciate the student’s contribution and wish them success in all future endeavors.", regular_font)
        ]

        # Function to wrap and draw text
        def draw_justified_paragraph(draw, paragraph, x, y, max_width, line_height):
            words_with_fonts = []
            for text, font in paragraph:
                for word in text.split():
                    words_with_fonts.append((word, font))

            lines = []
            current_line = []
            line_width = 0

            for word, font in words_with_fonts:
                word_width = draw.textlength(word, font=font)
                space_width = draw.textlength(' ', font=font) if current_line else 0
                if line_width + word_width + space_width <= max_width:
                    current_line.append((word, font))
                    line_width += word_width + space_width
                else:
                    lines.append(current_line)
                    current_line = [(word, font)]
                    line_width = word_width

            if current_line:
                lines.append(current_line)

            for i, line in enumerate(lines):
                total_text_width = sum(draw.textlength(word, font=font) for word, font in line)
                space_count = len(line) - 1
                if space_count > 0 and i < len(lines) - 1:
                    space_width = (max_width - total_text_width) / space_count
                else:
                    space_width = draw.textlength(' ', font=regular_font)

                dx = x
                for j, (word, font) in enumerate(line):
                    draw.text((dx, y), word, fill="black", font=font)
                    dx += draw.textlength(word, font=font)
                    if j != len(line) - 1:
                        dx += space_width
                y += line_height
            return y

        y = 880
        y = draw_justified_paragraph(draw, part1, x=200, y=y, max_width=1000, line_height=32)
        y += 20
        y = draw_justified_paragraph(draw, part2, x=200, y=y, max_width=1000, line_height=32)
        y += 20
        y = draw_justified_paragraph(draw, part3, x=200, y=y, max_width=1000, line_height=32)

        # Save to Excel
        excel_path = "certificates.xlsx"
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Certificates"
            ws.append([
                "Name", "ID No", "Department", "University", "Role",
                "Start Date", "End Date", "Issue Date", "Reg No"
            ])
        else:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

        ws.append([
            name, id_no, department, university, role,
            start_date, end_date, issue_date, reg_no
        ])

        # Auto adjust width
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(excel_path)

        # Save image
        img_io = io.BytesIO()
        format_map = {'png': 'PNG', 'jpg': 'JPEG', 'pdf': 'PDF'}
        image.save(img_io, format=format_map[file_format])
        img_io.seek(0)

        preview_cache['image'] = img_io.getvalue()
        preview_cache['format'] = file_format
        preview_cache['filename'] = f"{name.replace(' ', '_')}_certificate.{file_format}"

        return render_template("preview.html", filename=preview_cache['filename'])

    except Exception as e:
        return f"An error occurred: {str(e)}"

@app.route('/preview_image')
def preview_image():
    if 'image' in preview_cache:
        return send_file(io.BytesIO(preview_cache['image']),
                         mimetype=f"image/{preview_cache['format']}")
    return "No image preview available", 404

@app.route('/download')
def download():
    if 'image' in preview_cache:
        return send_file(io.BytesIO(preview_cache['image']),
                         mimetype=f"image/{preview_cache['format']}",
                         as_attachment=True,
                         download_name=preview_cache['filename'])
    return redirect(url_for('index'))

@app.route('/cancel')
def cancel():
    preview_cache.clear()
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
